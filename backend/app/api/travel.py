import json
import os
from datetime import date, datetime
from io import BytesIO
from typing import Any
from urllib.parse import urlencode
from urllib.request import urlopen
from uuid import uuid4

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from ..auth import current_user, require_writer
from ..database import get_db
from ..models import TravelDay, TravelImportJob, TravelLeg, TravelPlan, TravelStop, User
from ..utils.time import now_beijing
from .settings import amap_web_service_key
from ..schemas import (
    TravelDayIn,
    TravelDayOut,
    TravelGeocodeIn,
    TravelGeocodeOut,
    TravelImportErrorOut,
    TravelImportOut,
    TravelLegIn,
    TravelLegOut,
    TravelPlanIn,
    TravelPlanOut,
    TravelStopIn,
    TravelStopOut,
)

router = APIRouter()

DAY_COLORS = ["#e11d8a", "#7c3aed", "#f97316", "#db2777", "#0f766e", "#d97706", "#dc2626", "#2563eb", "#9333ea", "#f43f5e", "#0891b2", "#65a30d"]
STOP_TYPES = {"start", "end", "scenic", "restaurant", "hotel", "station", "airport", "shopping", "rest", "other"}
TRANSPORTS = {"walk", "bike", "drive", "taxi", "bus", "subway", "train", "flight", "other"}
TYPE_MAP = {
    "出发点": "start",
    "终点": "end",
    "景点": "scenic",
    "餐厅": "restaurant",
    "酒店": "hotel",
    "车站": "station",
    "机场": "airport",
    "购物": "shopping",
    "休息": "rest",
    "其他": "other",
}
TRANSPORT_MAP = {
    "步行": "walk",
    "骑行": "bike",
    "驾车": "drive",
    "出租车": "taxi",
    "打车": "taxi",
    "公交": "bus",
    "地铁": "subway",
    "火车": "train",
    "飞机": "flight",
    "其他": "other",
}
TYPE_LABELS = {value: key for key, value in TYPE_MAP.items()}
TRANSPORT_LABELS = {value: key for key, value in TRANSPORT_MAP.items()}


def stop_type_labels(value: str) -> str:
    return TYPE_LABELS.get(value, "其他")


def transport_labels(value: str) -> str:
    if value == "taxi":
        return "打车"
    return TRANSPORT_LABELS.get(value, "其他")


def parse_transports(raw: str | None, fallback: str = "other") -> list[str]:
    values: list[str] = []
    if raw:
        try:
            parsed = json.loads(raw)
            if isinstance(parsed, list):
                values = [normalize_transport(str(item)) for item in parsed]
        except json.JSONDecodeError:
            values = [normalize_transport(part) for part in raw.split(",")]
    if not values:
        values = [normalize_transport(fallback)]
    return [item for index, item in enumerate(values) if item in TRANSPORTS and item not in values[:index]]


def route_geometry_for(from_stop: TravelStop | None, to_stop: TravelStop | None) -> str:
    if not from_stop or not to_stop or from_stop.longitude is None or from_stop.latitude is None or to_stop.longitude is None or to_stop.latitude is None:
        return ""
    return json.dumps([[from_stop.longitude, from_stop.latitude], [to_stop.longitude, to_stop.latitude]], ensure_ascii=False)


def parse_route_geometry(raw: str) -> list[list[float]]:
    if not raw:
        return []
    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        return []


def parse_image_urls(raw: str | None, fallback: str | None = None) -> list[str]:
    urls: list[str] = []
    if raw:
        try:
            value = json.loads(raw)
            if isinstance(value, list):
                urls.extend(str(item) for item in value if item)
        except json.JSONDecodeError:
            urls.extend(part.strip() for part in raw.split(",") if part.strip())
    if fallback and fallback not in urls:
        urls.insert(0, fallback)
    return urls


def stop_out(stop: TravelStop, day_number: int) -> TravelStopOut:
    image_urls = parse_image_urls(stop.image_urls, stop.image_url)
    return TravelStopOut(
        id=stop.id,
        dayNumber=day_number,
        order=stop.sort_order,
        type=stop.stop_type,
        name=stop.name,
        city=stop.city,
        district=stop.district,
        address=stop.address,
        longitude=stop.longitude,
        latitude=stop.latitude,
        coordSystem=stop.coord_system,
        arriveTime=stop.arrive_time,
        leaveTime=stop.leave_time,
        stayMinutes=stop.stay_minutes,
        cost=stop.cost,
        openTime=stop.open_time,
        ticket=stop.ticket,
        needReservation=stop.need_reservation,
        mealType=stop.meal_type,
        recommendedFood=stop.recommended_food,
        guide=stop.guide,
        warning=stop.warning,
        note=stop.note,
        imageUrl=image_urls[0] if image_urls else stop.image_url,
        imageUrls=image_urls,
    )


def leg_out(leg: TravelLeg, day_number: int, stops_by_id: dict[str, TravelStop]) -> TravelLegOut:
    from_stop = stops_by_id.get(leg.from_stop_id or "")
    to_stop = stops_by_id.get(leg.to_stop_id or "")
    transports = parse_transports(leg.transports, leg.transport)
    return TravelLegOut(
        id=leg.id,
        dayNumber=day_number,
        fromOrder=from_stop.sort_order if from_stop else 0,
        toOrder=to_stop.sort_order if to_stop else 0,
        fromStopId=leg.from_stop_id,
        toStopId=leg.to_stop_id,
        transport=transports[0],
        transports=transports,
        departTime=leg.depart_time,
        arriveTime=leg.arrive_time,
        plannedMinutes=leg.planned_minutes,
        plannedCost=leg.planned_cost,
        plannedDistanceKm=leg.planned_distance_km,
        useMapRoute=leg.use_map_route,
        mapMinutes=leg.map_minutes,
        mapDistanceKm=leg.map_distance_km,
        routeGeometry=parse_route_geometry(leg.route_geometry),
        note=leg.note,
    )


def plan_out(plan: TravelPlan) -> TravelPlanOut:
    days: list[TravelDayOut] = []
    for day in sorted(plan.days, key=lambda item: item.day_number):
        stops = sorted(day.stops, key=lambda item: item.sort_order)
        stops_by_id = {stop.id: stop for stop in stops}
        days.append(
            TravelDayOut(
                id=day.id,
                dayNumber=day.day_number,
                date=day.date,
                title=day.title,
                summary=day.summary,
                themeColor=day.theme_color,
                stops=[stop_out(stop, day.day_number) for stop in stops],
                legs=[leg_out(leg, day.day_number, stops_by_id) for leg in sorted(day.legs, key=lambda item: (item.depart_time, item.id))],
            )
        )
    return TravelPlanOut(
        id=plan.id,
        countdownId=plan.countdown_id,
        travelCode=plan.travel_code,
        title=plan.title,
        destination=plan.destination,
        startDate=plan.start_date,
        endDate=plan.end_date,
        intro=plan.intro,
        coverUrl=plan.cover_url,
        defaultMapMode=plan.default_map_mode,
        createdBy=plan.created_by,
        updatedBy=plan.updated_by,
        createdAt=plan.created_at,
        updatedAt=plan.updated_at,
        days=days,
    )


def normalize_stop_type(value: str) -> str:
    normalized = TYPE_MAP.get(value.strip(), value.strip())
    return normalized if normalized in STOP_TYPES else "other"


def normalize_transport(value: str) -> str:
    normalized = TRANSPORT_MAP.get(value.strip(), value.strip())
    return normalized if normalized in TRANSPORTS else "other"


def normalize_transports(values: list[str] | None, fallback: str = "other") -> list[str]:
    normalized = [normalize_transport(str(value)) for value in (values or [])]
    normalized = [item for index, item in enumerate(normalized) if item in TRANSPORTS and item not in normalized[:index]]
    return normalized or [normalize_transport(fallback)]


def split_transport_text(value: str) -> list[str]:
    value = value.replace("，", "+").replace(",", "+").replace("/", "+").replace("、", "+")
    return [part.strip() for part in value.split("+") if part.strip()]


def geocode_address(city: str, district: str, address: str, key_override: str = "") -> TravelGeocodeOut:
    key = key_override or os.getenv("AMAP_WEB_SERVICE_KEY", "")
    keyword = " ".join(part for part in [city, district, address] if part).strip()
    if not key:
        return TravelGeocodeOut(ok=False, message="未配置 AMAP_WEB_SERVICE_KEY，无法自动定位")
    if not keyword:
        return TravelGeocodeOut(ok=False, message="地址为空，无法定位")
    query = urlencode({"key": key, "address": keyword, "city": city or ""})
    try:
        with urlopen(f"https://restapi.amap.com/v3/geocode/geo?{query}", timeout=10) as response:
            data = json.loads(response.read().decode("utf-8"))
    except Exception as exc:
        return TravelGeocodeOut(ok=False, message=f"地理编码请求失败：{exc}")
    geocodes = data.get("geocodes") or []
    if data.get("status") != "1" or not geocodes:
        return TravelGeocodeOut(ok=False, message=data.get("info") or "地址无法定位")
    location = geocodes[0].get("location", "")
    try:
        longitude, latitude = [float(part) for part in location.split(",", 1)]
    except ValueError:
        return TravelGeocodeOut(ok=False, message="地图服务返回的坐标格式不正确")
    return TravelGeocodeOut(ok=True, longitude=longitude, latitude=latitude, formattedAddress=geocodes[0].get("formatted_address", keyword))


def build_plan(db: Session, countdown_id: str, payload: TravelPlanIn, user: User, existing: TravelPlan | None = None, auto_geocode: bool = True) -> TravelPlan:
    now = now_beijing()
    plan = existing or TravelPlan(id=uuid4().hex, countdown_id=countdown_id, created_by=user.role, created_at=now)
    plan.travel_code = payload.travelCode
    plan.title = payload.title
    plan.destination = payload.destination
    plan.start_date = payload.startDate
    plan.end_date = payload.endDate
    plan.intro = payload.intro
    plan.cover_url = payload.coverUrl
    plan.default_map_mode = payload.defaultMapMode
    if existing:
        plan.updated_by = user.role
        plan.updated_at = now
        plan.days.clear()
    db.add(plan)
    db.flush()

    for index, day_payload in enumerate(sorted(payload.days, key=lambda item: item.dayNumber)):
        day = TravelDay(
            id=uuid4().hex,
            plan_id=plan.id,
            day_number=day_payload.dayNumber,
            date=day_payload.date,
            title=day_payload.title,
            summary=day_payload.summary,
            theme_color=day_payload.themeColor or DAY_COLORS[index % len(DAY_COLORS)],
        )
        db.add(day)
        db.flush()
        order_map: dict[int, TravelStop] = {}
        for stop_payload in sorted(day_payload.stops, key=lambda item: item.order):
            image_urls = [url for url in (stop_payload.imageUrls or []) if url]
            if stop_payload.imageUrl and stop_payload.imageUrl not in image_urls:
                image_urls.insert(0, stop_payload.imageUrl)
            longitude = stop_payload.longitude
            latitude = stop_payload.latitude
            locate_text = (stop_payload.address or stop_payload.city or "").strip()
            if auto_geocode and locate_text:
                geo = geocode_address(stop_payload.city, stop_payload.district, locate_text, amap_web_service_key(db))
                if geo.ok:
                    longitude = geo.longitude
                    latitude = geo.latitude
            stop = TravelStop(
                id=stop_payload.id or uuid4().hex,
                day_id=day.id,
                sort_order=stop_payload.order,
                stop_type=normalize_stop_type(stop_payload.type),
                name=stop_payload.name,
                city=stop_payload.city,
                district=stop_payload.district,
                address=stop_payload.address,
                longitude=longitude,
                latitude=latitude,
                coord_system=stop_payload.coordSystem,
                arrive_time=stop_payload.arriveTime,
                leave_time=stop_payload.leaveTime,
                stay_minutes=stop_payload.stayMinutes,
                cost=stop_payload.cost,
                open_time=stop_payload.openTime,
                ticket=stop_payload.ticket,
                need_reservation=stop_payload.needReservation,
                meal_type=stop_payload.mealType,
                recommended_food=stop_payload.recommendedFood,
                guide=stop_payload.guide,
                warning=stop_payload.warning,
                note=stop_payload.note,
                image_url=image_urls[0] if image_urls else stop_payload.imageUrl,
                image_urls=json.dumps(image_urls, ensure_ascii=False),
            )
            db.add(stop)
            order_map[stop.sort_order] = stop
        db.flush()

        legs = day_payload.legs or auto_legs(day_payload)
        for leg_payload in legs:
            from_stop = order_map.get(leg_payload.fromOrder)
            to_stop = order_map.get(leg_payload.toOrder)
            route_geometry = route_geometry_for(from_stop, to_stop)
            transports = normalize_transports(leg_payload.transports, leg_payload.transport)
            db.add(
                TravelLeg(
                    id=leg_payload.id or uuid4().hex,
                    day_id=day.id,
                    from_stop_id=from_stop.id if from_stop else None,
                    to_stop_id=to_stop.id if to_stop else None,
                    transport=transports[0],
                    transports=json.dumps(transports, ensure_ascii=False),
                    depart_time=leg_payload.departTime,
                    arrive_time=leg_payload.arriveTime,
                    planned_minutes=leg_payload.plannedMinutes,
                    planned_cost=leg_payload.plannedCost,
                    planned_distance_km=leg_payload.plannedDistanceKm,
                    use_map_route=leg_payload.useMapRoute,
                    map_minutes=leg_payload.mapMinutes,
                    map_distance_km=leg_payload.mapDistanceKm,
                    route_geometry=route_geometry,
                    note=leg_payload.note,
                )
            )
    return plan


def auto_legs(day: TravelDayIn) -> list[TravelLegIn]:
    stops = sorted(day.stops, key=lambda item: item.order)
    return [
        TravelLegIn(dayNumber=day.dayNumber, fromOrder=stops[index].order, toOrder=stops[index + 1].order, transport="other")
        for index in range(len(stops) - 1)
    ]


@router.get("/{countdown_id}/travel-plan", response_model=TravelPlanOut | None)
def get_travel_plan(countdown_id: str, db: Session = Depends(get_db)):
    plan = db.query(TravelPlan).filter(TravelPlan.countdown_id == countdown_id).first()
    return plan_out(plan) if plan else None


@router.get("/travel-plan/template")
def download_travel_template():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation

    workbook = Workbook()
    info = workbook.active
    info.title = "行程信息"
    info.append(["行程编号", "行程标题", "目的地", "开始日期", "结束日期", "行程简介", "封面图片", "默认地图模式"])
    info.append(["hangzhou-2026", "杭州三日浪漫旅行", "杭州", "2026-10-01", "2026-10-03", "一起去看秋天的西湖", "cover.jpg", "2D"])
    stops = workbook.create_sheet("行程节点")
    stops.append(["第几天", "顺序", "日期", "类型", "地点名称", "城市/区域", "详细地址", "到达时间", "离开时间", "停留分钟", "预计花费", "开放时间", "门票", "是否预约", "饭点类型", "推荐菜品", "攻略内容", "避雷提示", "备注", "图片文件名"])
    stops.append([1, 1, "2026-10-01", "车站", "杭州东站", "杭州", "杭州东站", "09:00", "09:20", 20, 0, "", "", "否", "", "", "到站后先买水", "节假日人多，提前出站", "", ""])
    stops.append([1, 2, "2026-10-01", "景点", "西湖断桥", "杭州", "北山街", "10:00", "11:30", 90, 0, "06:00-22:00", "免费", "否", "", "", "建议从断桥开始慢慢散步", "人多时不要在桥上停太久", "带一把伞", "xihu.jpg"])
    legs = workbook.create_sheet("交通段")
    legs.append(["第几天", "起点顺序", "终点顺序", "交通方式", "出发时间", "到达时间", "预计分钟", "预计花费", "预计距离公里", "是否使用地图路线", "交通备注"])
    legs.append([1, 1, 2, "地铁+步行", "09:20", "09:55", 35, 6, 8.5, "是", "从地铁 B 口出站"])
    tips = workbook.create_sheet("填写说明")
    tips.append(["说明"])
    tips.append(["地点不用填写经纬度，系统会根据“地点名称 + 城市/区域 + 详细地址”自动定位。"])
    tips.append(["交通方式支持多选写法，例如：地铁+步行、打车+步行。"])
    tips.append(["类型可选：出发点、终点、景点、餐厅、酒店、车站、机场、购物、休息、其他。"])
    tips.append(["时间请使用 09:30 这种 24 小时格式。"])
    header_fill = PatternFill("solid", fgColor="FFE1F0")
    for sheet in [info, stops, legs, tips]:
      for cell in sheet[1]:
          cell.font = Font(bold=True, color="8B3C6B")
          cell.fill = header_fill
      for column in sheet.columns:
          sheet.column_dimensions[column[0].column_letter].width = min(22, max(12, max(len(str(cell.value or "")) for cell in column) + 2))
    type_rule = DataValidation(type="list", formula1='"出发点,终点,景点,餐厅,酒店,车站,机场,购物,休息,其他"', allow_blank=False)
    stops.add_data_validation(type_rule)
    type_rule.add("D2:D300")
    transport_rule = DataValidation(type="list", formula1='"步行,骑行,驾车,打车,公交,地铁,火车,飞机,其他,地铁+步行,打车+步行"', allow_blank=True)
    legs.add_data_validation(transport_rule)
    transport_rule.add("D2:D300")
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=travel_plan_template.xlsx"},
    )


@router.get("/{countdown_id}/travel-plan/export")
def export_travel_plan(countdown_id: str, db: Session = Depends(get_db)):
    from openpyxl import Workbook

    plan = db.query(TravelPlan).filter(TravelPlan.countdown_id == countdown_id).first()
    if not plan:
        raise HTTPException(status_code=404, detail="还没有旅行攻略可以导出")

    workbook = Workbook()
    info = workbook.active
    info.title = "行程信息"
    info.append(["行程编号", "行程标题", "目的地", "开始日期", "结束日期", "行程简介", "封面图片", "默认地图模式"])
    info.append([
        plan.travel_code or plan.countdown_id,
        plan.title,
        plan.destination,
        plan.start_date.isoformat(),
        plan.end_date.isoformat(),
        plan.intro or "",
        plan.cover_url or "",
        "2D",
    ])

    stops_sheet = workbook.create_sheet("行程节点")
    stops_sheet.append(["第几天", "顺序", "日期", "类型", "地点名称", "城市/区域", "详细地址", "到达时间", "离开时间", "停留分钟", "预计花费", "开放时间", "门票", "是否预约", "饭点类型", "推荐菜品", "攻略内容", "避雷提示", "备注", "图片文件名"])
    legs_sheet = workbook.create_sheet("交通段")
    legs_sheet.append(["第几天", "起点顺序", "终点顺序", "交通方式", "出发时间", "到达时间", "预计分钟", "预计花费", "预计距离公里", "是否使用地图路线", "交通备注"])

    for day in sorted(plan.days, key=lambda item: item.day_number):
        stops = sorted(day.stops, key=lambda item: item.sort_order)
        stops_by_id = {stop.id: stop for stop in stops}
        for stop in stops:
            stops_sheet.append([
                day.day_number,
                stop.sort_order,
                day.date.isoformat(),
                stop_type_labels(stop.stop_type),
                stop.name,
                stop.city,
                stop.address,
                stop.arrive_time,
                stop.leave_time,
                stop.stay_minutes,
                stop.cost,
                stop.open_time,
                stop.ticket,
                "是" if stop.need_reservation else "否",
                stop.meal_type,
                stop.recommended_food,
                stop.guide,
                stop.warning,
                stop.note,
                ",".join(parse_image_urls(stop.image_urls, stop.image_url)),
            ])
        def leg_sort_key(item: TravelLeg) -> int:
            from_stop = stops_by_id.get(item.from_stop_id or "")
            return from_stop.sort_order if from_stop else 0

        for leg in sorted(day.legs, key=leg_sort_key):
            from_stop = stops_by_id.get(leg.from_stop_id or "")
            to_stop = stops_by_id.get(leg.to_stop_id or "")
            legs_sheet.append([
                day.day_number,
                from_stop.sort_order if from_stop else "",
                to_stop.sort_order if to_stop else "",
                "+".join(transport_labels(item) for item in parse_transports(leg.transports, leg.transport)),
                leg.depart_time,
                leg.arrive_time,
                leg.planned_minutes,
                leg.planned_cost,
                leg.planned_distance_km,
                "是" if leg.use_map_route else "否",
                leg.note,
            ])

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=travel_plan_export.xlsx"},
    )


@router.get("/{countdown_id}/travel-plan/map-data", response_model=TravelPlanOut | None)
def get_travel_map_data(countdown_id: str, db: Session = Depends(get_db)):
    return get_travel_plan(countdown_id, db)


@router.post("/{countdown_id}/travel-plan", response_model=TravelPlanOut)
def create_travel_plan(countdown_id: str, payload: TravelPlanIn, db: Session = Depends(get_db), user: User = Depends(require_writer)):
    existing = db.query(TravelPlan).filter(TravelPlan.countdown_id == countdown_id).first()
    if existing:
        raise HTTPException(status_code=409, detail="这个约定已经有旅行攻略了，请使用编辑或导入覆盖。")
    plan = build_plan(db, countdown_id, payload, user)
    db.commit()
    db.refresh(plan)
    return plan_out(plan)


@router.put("/{countdown_id}/travel-plan", response_model=TravelPlanOut)
def update_travel_plan(countdown_id: str, payload: TravelPlanIn, db: Session = Depends(get_db), user: User = Depends(require_writer)):
    existing = db.query(TravelPlan).filter(TravelPlan.countdown_id == countdown_id).first()
    plan = build_plan(db, countdown_id, payload, user, existing)
    db.commit()
    db.refresh(plan)
    return plan_out(plan)


@router.delete("/{countdown_id}/travel-plan")
def delete_travel_plan(countdown_id: str, db: Session = Depends(get_db), user: User = Depends(require_writer)):
    plan = db.query(TravelPlan).filter(TravelPlan.countdown_id == countdown_id).first()
    if plan:
        db.delete(plan)
        db.commit()
    return {"ok": True}


@router.post("/{countdown_id}/travel-plan/geocode", response_model=TravelGeocodeOut)
def geocode_travel_stop(countdown_id: str, payload: TravelGeocodeIn, db: Session = Depends(get_db), user: User | None = Depends(current_user)):
    return geocode_address(payload.city, payload.district, payload.address, amap_web_service_key(db))


@router.post("/{countdown_id}/travel-plan/generate-routes", response_model=TravelPlanOut | None)
def generate_routes(countdown_id: str, db: Session = Depends(get_db), user: User = Depends(require_writer)):
    plan = db.query(TravelPlan).filter(TravelPlan.countdown_id == countdown_id).first()
    if not plan:
        return None
    for day in plan.days:
        stops_by_id = {stop.id: stop for stop in day.stops}
        for leg in day.legs:
            leg.route_geometry = route_geometry_for(stops_by_id.get(leg.from_stop_id or ""), stops_by_id.get(leg.to_stop_id or ""))
    plan.updated_by = user.role
    plan.updated_at = now_beijing()
    db.commit()
    db.refresh(plan)
    return plan_out(plan)


def cell(row: dict[str, Any], name: str, default: Any = "") -> Any:
    value = row.get(name)
    return default if value is None else value


def as_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str) and value.strip():
        return date.fromisoformat(value.strip())
    return None


def as_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    return float(value)


def as_int(value: Any) -> int | None:
    if value in (None, ""):
        return None
    return int(value)


def as_bool(value: Any) -> bool:
    return str(value or "").strip() in {"是", "true", "True", "1", "yes", "Y"}


def rows_from_sheet(workbook, sheet_name: str) -> list[tuple[int, dict[str, Any]]]:
    if sheet_name not in workbook.sheetnames:
        return []
    sheet = workbook[sheet_name]
    headers = [str(value).strip() if value is not None else "" for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    rows: list[tuple[int, dict[str, Any]]] = []
    for row_index, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(value not in (None, "") for value in values):
            continue
        rows.append((row_index, {headers[index]: values[index] if index < len(values) else None for index in range(len(headers))}))
    return rows


def import_errors_to_models(errors: list[dict[str, Any]]) -> list[TravelImportErrorOut]:
    return [TravelImportErrorOut(**error) for error in errors]


def parse_excel_plan(content: bytes, web_service_key: str = "", auto_geocode: bool = False) -> tuple[TravelPlanIn | None, list[dict[str, Any]]]:
    from openpyxl import load_workbook

    errors: list[dict[str, Any]] = []
    workbook = load_workbook(BytesIO(content), data_only=True)
    info_rows = rows_from_sheet(workbook, "行程信息")
    stop_rows = rows_from_sheet(workbook, "行程节点")
    leg_rows = rows_from_sheet(workbook, "交通段")

    if not info_rows:
        errors.append({"sheet": "行程信息", "row": 1, "field": "行程信息", "reason": "缺少行程信息", "suggestion": "请至少填写一行行程信息。"})
        return None, errors
    _, info = info_rows[0]
    required_info = ["行程编号", "行程标题", "目的地", "开始日期", "结束日期"]
    for field in required_info:
        if not cell(info, field):
            errors.append({"sheet": "行程信息", "row": 2, "field": field, "reason": "必填字段为空", "suggestion": f"请填写 {field}。"})
    try:
        start_date = as_date(cell(info, "开始日期"))
        end_date = as_date(cell(info, "结束日期"))
    except Exception:
        start_date = end_date = None
        errors.append({"sheet": "行程信息", "row": 2, "field": "开始日期/结束日期", "reason": "日期格式不正确", "suggestion": "请使用 YYYY-MM-DD。"})
    if not start_date or not end_date or errors:
        return None, errors

    days: dict[int, TravelDayIn] = {}
    seen_orders: set[tuple[int, int]] = set()
    for row_index, row in stop_rows:
        try:
            day_number = int(cell(row, "第几天"))
            order = int(cell(row, "顺序"))
            stop_date = as_date(cell(row, "日期"))
            if not stop_date or stop_date < start_date or stop_date > end_date:
                errors.append({"sheet": "行程节点", "row": row_index, "field": "日期", "reason": "日期不在旅行范围内", "suggestion": "请确认日期位于开始日期和结束日期之间。"})
                continue
            if day_number < 1:
                errors.append({"sheet": "行程节点", "row": row_index, "field": "第几天", "reason": "第几天必须从 1 开始", "suggestion": "请填写 1 或更大的数字。"})
                continue
            if (day_number, order) in seen_orders:
                errors.append({"sheet": "行程节点", "row": row_index, "field": "顺序", "reason": "同一天顺序重复", "suggestion": "请调整当天地点顺序。"})
                continue
            seen_orders.add((day_number, order))
            stop_type = normalize_stop_type(str(cell(row, "类型")))
            if str(cell(row, "类型")) and stop_type == "other" and str(cell(row, "类型")).strip() not in {"其他", "other"}:
                errors.append({"sheet": "行程节点", "row": row_index, "field": "类型", "reason": "地点类型不在允许范围内", "suggestion": "可填写景点、餐厅、酒店、车站、机场、购物、休息、其他。"})
            longitude = as_float(cell(row, "经度"))
            latitude = as_float(cell(row, "纬度"))
            city = str(cell(row, "城市") or cell(row, "城市/区域"))
            district = str(cell(row, "区县"))
            address = str(cell(row, "详细地址") or cell(row, "地点名称"))
            if auto_geocode and (longitude is None or latitude is None):
                geo = geocode_address(city, district, address, web_service_key)
                if geo.ok:
                    longitude, latitude = geo.longitude, geo.latitude
                else:
                    errors.append({"sheet": "行程节点", "row": row_index, "field": "城市/区域 + 详细地址", "reason": geo.message, "suggestion": "请在网页里设置高德 Web 服务 Key 后重新导入，或先保存攻略再在编辑器里点“按地点定位”。"})
            stop = TravelStopIn(
                dayNumber=day_number,
                order=order,
                type=stop_type,
                name=str(cell(row, "地点名称")),
                city=city,
                district=district,
                address=address,
                longitude=longitude,
                latitude=latitude,
                coordSystem=str(cell(row, "坐标系", "gcj02") or "gcj02"),
                arriveTime=str(cell(row, "到达时间")),
                leaveTime=str(cell(row, "离开时间")),
                stayMinutes=as_int(cell(row, "停留分钟")),
                cost=as_float(cell(row, "预计花费")),
                openTime=str(cell(row, "开放时间")),
                ticket=str(cell(row, "门票")),
                needReservation=as_bool(cell(row, "是否预约")),
                mealType=str(cell(row, "饭点类型")),
                recommendedFood=str(cell(row, "推荐菜品")),
                guide=str(cell(row, "攻略内容")),
                warning=str(cell(row, "避雷提示")),
                note=str(cell(row, "备注")),
                imageUrl=(parse_image_urls(str(cell(row, "图片文件名")))[0] if parse_image_urls(str(cell(row, "图片文件名"))) else None),
                imageUrls=parse_image_urls(str(cell(row, "图片文件名"))),
            )
            days.setdefault(day_number, TravelDayIn(dayNumber=day_number, date=stop_date, title=f"Day {day_number}", themeColor=DAY_COLORS[(day_number - 1) % len(DAY_COLORS)])).stops.append(stop)
        except Exception as exc:
            errors.append({"sheet": "行程节点", "row": row_index, "field": "整行", "reason": str(exc), "suggestion": "请检查必填字段、数字和日期格式。"})

    for row_index, row in leg_rows:
        try:
            day_number = int(cell(row, "第几天"))
            transport_text = str(cell(row, "交通方式"))
            transports = normalize_transports(split_transport_text(transport_text), transport_text)
            transport = transports[0]
            if transport_text and transports == ["other"] and transport_text.strip() not in {"其他", "other"}:
                errors.append({"sheet": "交通段", "row": row_index, "field": "交通方式", "reason": "交通方式不在允许范围内", "suggestion": "可填写步行、骑行、驾车、打车、公交、地铁、火车、飞机、其他。"})
            if day_number not in days:
                errors.append({"sheet": "交通段", "row": row_index, "field": "第几天", "reason": "找不到对应行程节点", "suggestion": "请先在行程节点中填写这一天。"})
                continue
            leg = TravelLegIn(
                dayNumber=day_number,
                fromOrder=int(cell(row, "起点顺序")),
                toOrder=int(cell(row, "终点顺序")),
                transport=transport,
                transports=transports,
                departTime=str(cell(row, "出发时间")),
                arriveTime=str(cell(row, "到达时间")),
                plannedMinutes=as_int(cell(row, "预计分钟")),
                plannedCost=as_float(cell(row, "预计花费")),
                plannedDistanceKm=as_float(cell(row, "预计距离公里")),
                useMapRoute=as_bool(cell(row, "是否使用地图路线", "是")),
                note=str(cell(row, "交通备注")),
            )
            orders = {stop.order for stop in days[day_number].stops}
            if leg.fromOrder not in orders or leg.toOrder not in orders:
                errors.append({"sheet": "交通段", "row": row_index, "field": "起点顺序/终点顺序", "reason": "起点或终点不存在", "suggestion": "请确认顺序在行程节点中存在。"})
                continue
            days[day_number].legs.append(leg)
        except Exception as exc:
            errors.append({"sheet": "交通段", "row": row_index, "field": "整行", "reason": str(exc), "suggestion": "请检查起点顺序、终点顺序和交通方式。"})

    plan = TravelPlanIn(
        travelCode=str(cell(info, "行程编号")),
        title=str(cell(info, "行程标题")),
        destination=str(cell(info, "目的地")),
        startDate=start_date,
        endDate=end_date,
        intro=str(cell(info, "行程简介")),
        coverUrl=str(cell(info, "封面图片")) or None,
        defaultMapMode=str(cell(info, "默认地图模式", "2D") or "2D"),
        days=[days[key] for key in sorted(days)],
    )
    return plan, errors


@router.post("/{countdown_id}/travel-plan/import", response_model=TravelImportOut)
async def import_travel_plan(
    countdown_id: str,
    file: UploadFile = File(...),
    strategy: str = Form("overwrite"),
    db: Session = Depends(get_db),
    user: User = Depends(require_writer),
):
    content = await file.read()
    job = TravelImportJob(id=uuid4().hex, countdown_id=countdown_id, status="running", created_by=user.role)
    db.add(job)
    db.commit()
    try:
        payload, errors = parse_excel_plan(content, amap_web_service_key(db), auto_geocode=False)
    except Exception as exc:
        errors = [{"sheet": "导入文件", "row": 0, "field": "Excel", "reason": str(exc), "suggestion": "请确认上传的是 travel_plan_template.xlsx 格式。"}]
        payload = None
    if errors or not payload:
        job.status = "failed"
        job.message = "导入校验失败"
        job.error_report = json.dumps(errors, ensure_ascii=False)
        db.commit()
        return TravelImportOut(ok=False, errors=import_errors_to_models(errors), message="导入校验失败，请按错误提示修改 Excel。")

    existing = db.query(TravelPlan).filter(TravelPlan.countdown_id == countdown_id).first()
    if existing and strategy == "cancel":
        return TravelImportOut(ok=False, message="已存在旅行攻略，已取消导入。")
    target_countdown_id = f"{countdown_id}-{uuid4().hex[:6]}" if existing and strategy == "copy" else countdown_id
    target_existing = None if strategy == "copy" else existing
    plan = build_plan(db, target_countdown_id, payload, user, target_existing, auto_geocode=False)
    job.status = "success"
    job.message = "导入完成"
    db.commit()
    db.refresh(plan)
    return TravelImportOut(ok=True, plan=plan_out(plan), message="旅行攻略已保存。没有坐标的地点可以稍后在编辑器里点“按地点定位”。")
